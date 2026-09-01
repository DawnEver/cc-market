"""The audit workbook: a reader who has not seen the pipeline must be able to use it.

Which means the sheet may not carry pipeline vocabulary, and every rule it states
has to be the rule the run actually applied — so the thresholds and the
restricted-country list are read from the run and from the policy, never frozen
into the script.
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest

pytest.importorskip("openpyxl", reason="the workbook script is behind the 'xlsx' extra")

import openpyxl


def _load():
    path = Path(__file__).resolve().parents[1] / "scripts" / "audit_xlsx.py"
    spec = importlib.util.spec_from_file_location("audit_xlsx", path)
    module = importlib.util.module_from_spec(spec)
    sys.modules["audit_xlsx"] = module
    spec.loader.exec_module(module)
    return module


@pytest.fixture()
def script():
    module = _load()
    yield module
    sys.modules.pop("audit_xlsx", None)


HEADER = [
    "rank",
    "reviewer",
    "person_id",
    "email",
    "institution",
    "current_country",
    "selected_for_contact",
    "filter_coi",
    "filter_coi_severity",
    "filter_banned_country",
    "filter_banned_country_is_banned",
    "filter_related_journal_publications",
    "filter_related_journal_count",
    "filter_related_journal_minimum",
    "filter_details",
]


def row(country: str, banned: str, coi: str, journals: str) -> list[str]:
    return [
        "1",
        "A Reviewer",
        "person-1",
        "a@uni.edu",
        "Some Uni",
        country,
        "False",
        coi,
        "0" if coi == "CLEAR" else "2",
        "FILTERED" if banned == "1" else "PASS",
        banned,
        "PASS" if int(journals) >= 3 else "FILTERED",
        journals,
        "3",
        "a sentence per rule",
    ]


def write_csv(tmp_path: Path, *rows: list[str], slug: str = "tte") -> Path:
    case = tmp_path / "ongoing" / f"{slug}-case"
    (case / "1-manuscript").mkdir(parents=True)
    (case / "1-manuscript" / "paper_profile.json").write_text(
        f'{{"journal": "{slug}"}}', encoding="utf-8"
    )
    shortlist = case / "5-shortlist"
    shortlist.mkdir()
    src = shortlist / "contact-list-audit.csv"
    lines = [",".join(HEADER)] + [",".join(r) for r in rows]
    src.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    return src


def test_the_restricted_list_comes_from_the_journal_policy(script, tmp_path):
    src = write_csv(tmp_path, row("IN", "1", "CLEAR", "4"), row("CN", "0", "CLEAR", "4"))

    script.build(src, src.with_suffix(".xlsx"))
    heading = [c.value for c in openpyxl.load_workbook(src.with_suffix(".xlsx"))["decision"][1]]

    # Iran restricts nobody in this file; the heading still names it, because the
    # heading states the rule rather than what happened to fire.
    assert "Rule: not working in India (IN) or Iran (IR)" in heading


def test_a_csv_that_contradicts_the_policy_is_refused(script, tmp_path):
    # China is not restricted for TTE, so a row flagged as restricted means the
    # policy has moved since the run and the workbook would state the wrong rule.
    src = write_csv(tmp_path, row("CN", "1", "CLEAR", "4"))

    with pytest.raises(SystemExit, match="disagrees with the policy"):
        script.build(src, src.with_suffix(".xlsx"))


def test_the_threshold_in_a_heading_is_read_from_the_run(script, tmp_path):
    src = write_csv(tmp_path, row("CN", "0", "CLEAR", "4"))
    raw = src.read_text(encoding="utf-8-sig").replace(",3,a sentence", ",5,a sentence")
    src.write_text(raw, encoding="utf-8-sig")

    script.build(src, src.with_suffix(".xlsx"))
    heading = [c.value for c in openpyxl.load_workbook(src.with_suffix(".xlsx"))["decision"][1]]

    assert "Rule: related journal papers ≥ 5" in heading


def test_a_heading_never_renders_an_unresolved_placeholder(script, tmp_path):
    src = write_csv(tmp_path, row("CN", "0", "CLEAR", "4"), row("CN", "0", "CLEAR", "2"))
    # Two different minima: the column is no longer a constant, so no threshold
    # can be quoted and the run must stop rather than print "{...}" in a heading.
    raw = src.read_text(encoding="utf-8-sig").splitlines()
    raw[2] = raw[2].replace(",3,a sentence", ",4,a sentence")
    src.write_text("\n".join(raw) + "\n", encoding="utf-8-sig")

    with pytest.raises(SystemExit, match="related_journal_minimum"):
        script.build(src, src.with_suffix(".xlsx"))


def test_no_pipeline_vocabulary_reaches_the_workbook(script, tmp_path):
    src = write_csv(tmp_path, row("IN", "1", "CLEAR", "4"), row("CN", "0", "FILTERED", "1"))

    script.build(src, src.with_suffix(".xlsx"))
    book = openpyxl.load_workbook(src.with_suffix(".xlsx"))
    seen = {
        str(cell.value)
        for sheet in book
        for line in sheet.iter_rows()
        for cell in line
        if cell.value is not None
    }

    assert not seen & {"PASS", "FILTERED", "VERIFY", "PREFERENCE_MISSED", "CLEAR"}
    assert not any(text.startswith("filter_") for text in seen)
    assert "person_id" not in seen


def test_a_quantifiable_rule_shows_its_number_and_the_verdict_only_colours_it(script, tmp_path):
    src = write_csv(tmp_path, row("CN", "0", "CLEAR", "4"), row("CN", "0", "CLEAR", "1"))

    script.build(src, src.with_suffix(".xlsx"))
    sheet = openpyxl.load_workbook(src.with_suffix(".xlsx"))["decision"]
    heading = [c.value for c in sheet[1]]
    at = heading.index("Rule: related journal papers ≥ 3")
    cells = [line[at] for line in sheet.iter_rows(min_row=2)]

    assert [cell.value for cell in cells] == [4, 1]
    assert cells[0].font.color.rgb.endswith("0B6B2E")
    assert cells[1].font.color.rgb.endswith("9C0006")


def test_why_not_recommended_is_a_sentence_not_a_field_name(script, tmp_path):
    src = write_csv(tmp_path, row("IN", "1", "CLEAR", "4"))

    script.build(src, src.with_suffix(".xlsx"))
    sheet = openpyxl.load_workbook(src.with_suffix(".xlsx"))["decision"]
    at = [c.value for c in sheet[1]].index("Why not recommended")

    assert sheet.cell(row=2, column=at + 1).value == "Works in a restricted country"
