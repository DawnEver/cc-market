#!/usr/bin/env python3
"""Turn a reviewer-discovery contact-list-audit.csv into a readable workbook.

The CSV is one row per candidate and one column per filter *input*: the right
shape for an audit trail, the wrong shape for a reader who does not already know
the pipeline. Out of sixty-odd columns only eight are conclusions; the rest are
the arithmetic behind them. So the workbook splits into three layers:

* ``decision`` — what an editor actually filters on: one verdict per dimension,
  a single ``blocking_reason``, and the contact details. Opens first.
* ``audit``    — every input, ``filter_`` prefixes dropped and the constant
  thresholds moved into the glossary, so what is left varies per person.
* ``columns``  — how to read the thing, then a line per column.

    uv run --extra xlsx python scripts/audit_xlsx.py \
        <workspace>/ongoing/<slug>/5-shortlist/contact-list-audit.csv

Writes <name>.xlsx beside the input, so it lands in the case workspace while
the code stays here. Nothing in this file is confidential: it reads whatever CSV
it is handed and never touches the manuscript.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DROP = ("person_id",)
RENAME = {"selected_for_contact": "recommend_for_reviewer", "filter_details": "reasoning"}

#: Columns that state a policy threshold rather than a fact about the person.
#: Dropped from the sheet when they hold one value for everybody — sixty columns
#: is already too many to scan, and a column that never varies cannot filter.
THRESHOLD_SUFFIXES = ("_minimum", "_maximum", "_target", "_window_years")

#: Spelling out a country reads better than a bare code in a column heading.
#: Display only — the list of restricted countries itself comes from the policy,
#: never from here, and a code absent from this map simply prints as the code.
COUNTRY_NAMES = {
    "CN": "China",
    "IN": "India",
    "IR": "Iran",
    "KP": "North Korea",
    "RU": "Russia",
    "SY": "Syria",
}

WHITE_COLUMNS = ("rank", "reviewer", "email", "institution")

#: Dimension boundaries, in sheet order. Used for the rule between blocks and to
#: group the glossary; the header itself stays grey so the eye rests on the name.
DIMENSIONS: list[tuple[str, tuple[str, ...]]] = [
    ("Identity", ("rank", "reviewer", "email", "institution", "current_country")),
    ("Decision", ("recommend_for_reviewer", "blocking_reason")),
    ("Conflict of interest", ("coi",)),
    ("Geography", ("author_country",)),
    ("Restricted country", ("banned_country",)),
    ("Related-journal record", ("related_",)),
    ("Recent activity", ("recent_activity", "recent_paper", "latest_year")),
    ("Doctoral floor", ("doctoral_year", "is_doctoral")),
    ("Invitation response", ("invitation_", "recent_invitation")),
    ("Unresponsive veteran", ("unresponsive_veteran", "veteran_")),
    ("Reasoning", ("reasoning",)),
]

#: What each column is called in the sheet. The internal name stays the key
#: everywhere in this script but never reaches the workbook: a reader who has
#: not seen the pipeline should not have to decode a field name to use the file.
LABELS: dict[str, str] = {
    "rank": "Rank",
    "reviewer": "Reviewer name",
    "email": "Email address",
    "institution": "Current institution",
    "current_country": "Current country",
    "recommend_for_reviewer": "Recommend as reviewer",
    "blocking_reason": "Why not recommended",
    "coi": "Rule: no conflict of interest (severity = 0)",
    "coi_severity": "Conflict severity (0 none, 1 review, 2 blocking)",
    "coi_finding_count": "Number of conflicts found",
    "author_country_reference": "Country compared with the submission",
    "author_country_known": "Is the current country known?",
    "author_country_cross_region": "In a different country from the submission? (adds score, never excludes)",
    "author_country_origin_count": "Countries the submission comes from",
    "banned_country": "Rule: not working in {restricted_countries}",
    "banned_country_known": "Is the country known, for the restricted-list check?",
    "banned_country_is_banned": "Working in {restricted_countries}",
    "related_journal_publications": "Rule: related journal papers ≥ {related_journal_minimum}",
    "related_journal_count": "Related journal papers — {related_journal_minimum} required",
    "related_journal_gap": "Related journal papers above the {related_journal_minimum} required",
    "related_journal_target_ratio": "Related journal papers ÷ target of {related_journal_target}",
    "related_nonjournal_count": "Related conference papers (never count towards the {related_journal_minimum})",
    "related_unknown_type_count": "Related papers, journal or conference unclear",
    "related_first_author_count": "Of those, as first author",
    "related_second_author_count": "Of those, as second author",
    "related_last_author_count": "Of those, as last author (the supervisor slot)",
    "related_middle_author_count": "Of those, as middle author",
    "related_corresponding_observed_count": "Of those, as corresponding author (seen on the paper itself)",
    "related_leadership_count": "Of those, in a leading role (first, last or corresponding)",
    "related_position_weight_sum": "Author-position score, total",
    "related_position_weight_mean": "Author-position score, average (1.0 = always leading)",
    "recent_activity": "Rule: ≥ {recent_paper_minimum} paper in the last {recent_window_years} years",
    "recent_activity_known": "Is a publication profile available?",
    "recent_paper_count": "Papers in the last {recent_window_years} years — {recent_paper_minimum} required",
    "latest_year": "Most recent publication year",
    "recent_paper_gap": "Papers above the {recent_paper_minimum} required",
    "doctoral_year": "Rule: if a PhD student, year of study ≥ {doctoral_year_minimum}",
    "is_doctoral": "Is a PhD student",
    "doctoral_year_known": "Is the year of PhD study known?",
    "doctoral_year_value": "Year of PhD study — {doctoral_year_minimum} required",
    "doctoral_year_gap": "PhD years above the {doctoral_year_minimum} required",
    "invitation_response": "Rule: answered ≥ {invitation_response_rate_minimum} of invitations in the last {invitation_window_years} years",
    "invitation_response_known": "Enough invitation history to judge?",
    "recent_invitation_count": "Invitations in the last {invitation_window_years} years — {recent_invitation_minimum} needed before judging",
    "invitation_response_rate": "Share of those invitations answered — {invitation_response_rate_minimum} required",
    "invitation_response_rate_gap": "Answer rate above the {invitation_response_rate_minimum} required",
    "unresponsive_veteran": "Rule: not (career ≥ {veteran_career_minimum} years and answered ≤ {veteran_response_rate_maximum} of ≥ {veteran_invitation_minimum} invitations)",
    "veteran_career_known": "Is the first publication year known?",
    "veteran_career_years": "Years since first publication — rule applies from {veteran_career_minimum}",
    "veteran_invitation_count": "Invitations ever received — {veteran_invitation_minimum} needed before judging",
    "veteran_response_rate": "Share answered over the whole career — unresponsive at {veteran_response_rate_maximum}",
    "reasoning": "Reason for each check, in words",
    "related_journal_minimum": "Related journal papers required",
    "related_journal_target": "Related journal papers aimed for",
    "recent_paper_minimum": "Papers required in the recent window",
    "recent_window_years": "Length of the recent-activity window, in years",
    "doctoral_year_minimum": "Minimum year of PhD study",
    "recent_invitation_minimum": "Invitations needed before the answer rate is judged",
    "invitation_response_rate_minimum": "Answer rate required",
    "invitation_window_years": "Length of the invitation window, in years",
    "veteran_career_minimum": "Career length that makes the veteran rule apply",
    "veteran_invitation_minimum": "Invitations needed before a veteran is judged",
    "veteran_response_rate_maximum": "Answer rate at or below which someone counts as unresponsive",
}


#: What ``Why not recommended`` says. A dimension name is not a reason.
BLOCKING_REASONS = {
    "coi": "Conflict of interest with the authors",
    "banned_country": "Works in a restricted country",
    "related_journal_publications": "Too few papers in related journals",
    "recent_activity": "Not publishing recently",
    "doctoral_year": "PhD student below the year floor",
    "invitation_response": "Rarely answers review invitations",
    "unresponsive_veteran": "Long career, no longer answers invitations",
}



#: Policy thresholds read out of the CSV, so a header states the rule this run
#: actually applied rather than a number frozen into this script.
THRESHOLDS: dict[str, str] = {}

#: The same thresholds unformatted, for comparing numbers against.
THRESHOLD_NUMBERS: dict[str, str] = {}


def _readable(name: str, value: str) -> str:
    """Rates read as percentages; everything else is already a plain number."""
    if "_rate_" in name:
        try:
            return f"{float(value):.0%}"
        except ValueError:
            pass
    return value


def label_of(name: str) -> str:
    return LABELS.get(name, name).format_map(THRESHOLDS)


def missing_thresholds(header: list[str]) -> list[str]:
    """Placeholders the headings need but the CSV did not supply.

    Only the columns actually present are checked. A journal that switches a
    rule off drops its columns from the CSV, and demanding a threshold for a
    rule nobody ran would refuse a perfectly good file.
    """
    import string

    wanted: set[str] = set()
    for name in header:
        for _, field, _, _ in string.Formatter().parse(LABELS.get(name, "")):
            if field:
                wanted.add(field)
    return sorted(wanted - set(THRESHOLDS))

#: The verdict column of each dimension, in the order a blocking reason is read.
VERDICTS_IN_ORDER = (
    "coi",
    "banned_country",
    "related_journal_publications",
    "recent_activity",
    "doctoral_year",
    "invitation_response",
    "unresponsive_veteran",
)

#: PASS / FILTERED / PREFERENCE_MISSED are pipeline vocabulary. A reader who has
#: never seen the pipeline gets the same states in words they already know.
VERDICT_TEXT = {
    "PASS": "Pass",
    "CLEAR": "Pass",
    "FILTERED": "Fail",
    "BLOCK": "Fail",
    "VERIFY": "No evidence",
    "REVIEW": "Check by hand",
    "PREFERENCE_MISSED": "Below preference",
}

VERDICT_STYLE = {
    "Pass": ("C6EFCE", "0B6B2E"),
    "Fail": ("F8C9C6", "9C0006"),
    "No evidence": ("FFE9A8", "8A6100"),
    "Below preference": ("FFE9A8", "8A6100"),
    "Check by hand": ("FFE9A8", "8A6100"),
}

#: A measured number is judged against the same threshold its rule uses, so the
#: figure itself reads pass or fail without cross-referencing the rule column.
#: (column, threshold column, the number must reach the threshold)
MEASURES: tuple[tuple[str, str, bool], ...] = (
    ("related_journal_count", "related_journal_minimum", True),
    ("recent_paper_count", "recent_paper_minimum", True),
    ("doctoral_year_value", "doctoral_year_minimum", True),
    ("invitation_response_rate", "invitation_response_rate_minimum", True),
    ("recent_invitation_count", "recent_invitation_minimum", True),
    ("veteran_invitation_count", "veteran_invitation_minimum", True),
)

#: A rule that measures something shows the measurement, not a verdict: the
#: heading already states the threshold, so the figure carries the whole story
#: and an editor can sort on it. The verdict still sets the colour, because it
#: is the only thing that knows the difference between failing a rule and having
#: nothing to judge. Rules with nothing to count — a country is on a list or it
#: is not — keep their word.
RULE_MEASURES = {
    "coi": "coi_severity",
    "related_journal_publications": "related_journal_count",
    "recent_activity": "recent_paper_count",
    "doctoral_year": "doctoral_year_value",
    "invitation_response": "invitation_response_rate",
}

#: Gap columns are already measured-minus-required, so zero is the line.
GAP_COLUMNS = (
    "related_journal_gap",
    "recent_paper_gap",
    "doctoral_year_gap",
    "invitation_response_rate_gap",
)

DECISION_COLUMNS = (
    "rank",
    "reviewer",
    "institution",
    "current_country",
    "email",
    "recommend_for_reviewer",
    "blocking_reason",
    *VERDICTS_IN_ORDER,
    "reasoning",
)

HOW_TO_READ = [
    ("How to read this workbook", ""),
    (
        "1. Start on the decision sheet.",
        "One row per candidate, one column per dimension. Filter ‘Recommend as reviewer’ = TRUE for the people who cleared everything.",
    ),
    (
        "2. To ask why somebody is out, filter ‘Why not recommended’.",
        "It states the first check they failed, and is blank for everyone still in play.",
    ),
    (
        "3. A Rule column shows a number where there is one to show.",
        "The heading states the threshold, so the figure underneath answers the rule on its own and can be sorted. Its colour is the verdict: green the rule is satisfied, red the candidate is excluded by it, amber either nothing on record to judge (the rule abstains rather than guess) or a preference rather than a requirement (it excludes nobody). Where there is nothing to count, or nothing was measured for this person, the cell says Pass, Fail, No evidence or Below preference instead.",
    ),
    (
        "4. The audit sheet is the arithmetic behind each verdict.",
        "Same rows, same order, every input the rules read. Go there to disagree with a verdict, not to make a shortlist.",
    ),
    (
        "5. Measured numbers are coloured against their own threshold.",
        "A count or rate is green once it reaches the figure named in its own column heading and red until it does, so a number reads pass or fail without looking across at the rule column.",
    ),
    ("", ""),
]


def dimension_of(name: str) -> str:
    best, best_len = "", -1
    for label, prefixes in DIMENSIONS:
        for prefix in prefixes:
            if name.startswith(prefix) and len(prefix) > best_len:
                best, best_len = label, len(prefix)
    return best


def cast(value: str):
    if value == "":
        return None
    if value in ("True", "False"):
        return value == "True"
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


GLOSSARY: dict[str, str] = {
    "rank": "Shortlist position, best expertise score first.",
    "reviewer": "Candidate name as resolved during identity matching.",
    "email": "Best verified contact address; details/evidence records where it came from.",
    "institution": "Current affiliation.",
    "current_country": "ISO code of the current affiliation, never nationality.",
    "recommend_for_reviewer": "TRUE only when every required dimension passed. This is the shortlist.",
    "blocking_reason": "The first dimension that excluded this candidate; blank when none did. Derived here, not in the CSV.",
    "coi": "Conflict-of-interest verdict under coi.toml plus the journal overlay: CLEAR or FILTERED.",
    "coi_severity": "0 clear, 1 review-level, 2 blocking.",
    "coi_finding_count": "Number of COI findings recorded against this person.",
    "author_country_reference": "The candidate's country read against the submission's country of origin.",
    "author_country_known": "1 when the current affiliation country is established.",
    "author_country_cross_region": "1 when the candidate sits outside the submission's origin country. Policy geo mode is prefer_cross_region, so this only scores; it never excludes.",
    "author_country_origin_count": "How many distinct origin countries the submission has.",
    "banned_country": "Fails anyone whose current affiliation is in a restricted country. The list is India (IN) and Iran (IR); an unknown country abstains rather than guesses.",
    "banned_country_known": "1 when the country is established; 0 forces VERIFY rather than a guess.",
    "banned_country_is_banned": "1 when the current affiliation country is on the list.",
    "related_journal_publications": "PASS once the verified related-journal count reaches the minimum.",
    "related_journal_count": "Verified journal papers in venues related to the submission.",
    "related_journal_gap": "count minus the minimum; negative means short of the floor.",
    "related_journal_target_ratio": "count divided by the target, 0..1.",
    "related_nonjournal_count": "Related conference or other non-journal items. They never satisfy the floor.",
    "related_unknown_type_count": "Related items whose venue type could not be resolved.",
    "related_first_author_count": "Related papers where the candidate is first author.",
    "related_second_author_count": "Related papers where the candidate is second author.",
    "related_last_author_count": "Related papers where the candidate is last author, the usual supervisor slot.",
    "related_middle_author_count": "Related papers where the candidate is a middle author.",
    "related_corresponding_observed_count": "Related papers where corresponding authorship was actually observed on the PDF or publisher page, never assumed.",
    "related_leadership_count": "Related papers in a leading role: first, last or corresponding.",
    "related_position_weight_sum": "Sum of authorship-position weights across the related papers.",
    "related_position_weight_mean": "Mean position weight. Near 1.0 means consistently leading.",
    "recent_activity": "PASS, or PREFERENCE_MISSED because this rule is set to prefer rather than require.",
    "recent_activity_known": "1 when a publication profile was available to read.",
    "recent_paper_count": "Papers inside the activity window.",
    "latest_year": "Year of the most recent publication.",
    "recent_paper_gap": "count minus the minimum.",
    "doctoral_year": "PASS unless the candidate is a doctoral student below the journal's floor.",
    "is_doctoral": "1 when the resolved rank is doctoral student.",
    "doctoral_year_known": "1 when the year of study is stated.",
    "doctoral_year_value": "Year of doctoral study; blank for everyone who is not a student.",
    "doctoral_year_gap": "value minus the minimum.",
    "invitation_response": "PASS / VERIFY. VERIFY means no invitation history exists, so the rule abstains instead of failing anyone.",
    "invitation_response_known": "1 when enough resolved invitations exist to judge.",
    "recent_invitation_count": "Resolved invitations inside the invitation window.",
    "invitation_response_rate": "Share answered, 0..1; blank when unknown.",
    "invitation_response_rate_gap": "rate minus the minimum.",
    "unresponsive_veteran": "PASS / VERIFY. Fires only on a long career AND a record of silence; VERIFY means the invitation record is too thin to judge.",
    "veteran_career_known": "1 when a first publication year was found.",
    "veteran_career_years": "Years since the first publication.",
    "veteran_invitation_count": "All resolved invitations on record, unwindowed.",
    "veteran_response_rate": "Lifetime response rate; blank when unknown.",
    "reasoning": "One human-readable sentence per rule. The reason an editor can disagree with.",
}


def _style_header(ws, header: list[str]) -> None:
    grey = PatternFill("solid", fgColor="D9D9D9")
    rule = Side(style="medium", color="8C97A3")
    thin = Side(style="thin", color="9AA5B1")
    previous = None
    for column, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=column)
        cell.value = label_of(name)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if name not in WHITE_COLUMNS:
            cell.fill = grey
        label = dimension_of(name)
        starts_block = previous is not None and label != previous
        cell.border = Border(bottom=thin, left=rule if starts_block else None)
        if starts_block:
            # The rule runs the full height so the blocks stay legible when the
            # header scrolls out of sight.
            for below in ws[get_column_letter(column)][1:]:
                below.border = Border(left=rule)
        previous = label


PASS_FONT = Font(color="0B6B2E", bold=True)
FAIL_FONT = Font(color="9C0006", bold=True)


def _floor_of(name: str) -> float | None:
    for column, threshold, _ in MEASURES:
        if column == name:
            try:
                return float(THRESHOLD_NUMBERS[threshold])
            except (KeyError, ValueError):
                return None
    return 0.0 if name in GAP_COLUMNS else None


def _style_body(ws, header: list[str], verdicts: list[list[str | None]]) -> None:
    floors = [_floor_of(name) for name in header]
    for index, row in enumerate(ws.iter_rows(min_row=2)):
        for position, cell in enumerate(row):
            verdict = verdicts[index][position]
            style = VERDICT_STYLE.get(verdict or cell.value) if isinstance(verdict or cell.value, str) else None
            if style:
                fill, colour = style
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(color=colour, bold=True)
            elif cell.value is True:
                cell.font = PASS_FONT
            elif cell.value is False:
                cell.font = FAIL_FONT
            elif floors[position] is not None and isinstance(cell.value, (int, float)):
                cell.font = PASS_FONT if cell.value >= floors[position] else FAIL_FONT


def _finish(
    ws,
    header: list[str],
    body: list[list],
    verdicts: list[list[str | None]],
    *,
    freeze: str,
    wide: tuple[str, ...],
) -> None:
    _style_header(ws, header)
    _style_body(ws, header, verdicts)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 58
    for column, name in enumerate(header, start=1):
        letter = get_column_letter(column)
        if name in wide:
            ws.column_dimensions[letter].width = 60
            for cell in ws[letter][1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            continue
        # The heading wraps, so it does not get a vote on the width: sizing to a
        # sixty-character rule sentence would push every neighbour off-screen.
        # The data decides, within a range narrow enough to keep the dimensions
        # readable side by side.
        widest = max([len(str(r[column - 1])) for r in body] or [0])
        ws.column_dimensions[letter].width = min(max(widest + 2, 8), 24)


def journal_of(src: Path) -> str:
    """The journal slug this case was scored under, from its own profile.

    Read from the derived profile the pipeline writes, not from the manuscript:
    one key, and nothing else in that file is touched. A workspace laid out some
    other way simply yields no slug, and the caller falls back.
    """
    for directory in src.parents:
        profile = directory / "1-manuscript" / "paper_profile.json"
        if profile.exists():
            try:
                return str(json.loads(profile.read_text(encoding="utf-8")).get("journal") or "")
            except (OSError, ValueError):
                return ""
    return ""


def restricted_countries(src: Path, journal: str) -> frozenset[str]:
    """The journal's restricted-country list, straight from the policy.

    Duplicating the list here would mean a workbook could go on naming countries
    a journal had stopped restricting. When the policy cannot be read at all —
    the library is not importable, or the slug is unknown — say so rather than
    guess a list.
    """
    slug = journal or journal_of(src)
    if not slug:
        return frozenset()
    try:
        from academia.reviewer.policy import load_policy
    except ImportError:  # pragma: no cover - only when run outside the project
        return frozenset()
    return load_policy(slug).restricted_country.upper_set("countries")


def describe_countries(codes: frozenset[str]) -> str:
    if not codes:
        return "a restricted country"
    named = sorted(f"{COUNTRY_NAMES.get(code, code)} ({code})" for code in codes)
    return " or ".join(named)


def build(src: Path, dst: Path, journal: str = "") -> tuple[int, int, int]:
    with src.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    raw_header, raw_body = rows[0], rows[1:]

    header, body_columns = [], []
    for index, name in enumerate(raw_header):
        if name in DROP:
            continue
        clean = RENAME.get(name, name.removeprefix("filter_"))
        header.append(clean)
        body_columns.append([row[index] for row in raw_body])

    # A threshold that is identical for everybody belongs in the glossary.
    constants: list[tuple[str, str]] = []
    keep = []
    for position, name in enumerate(header):
        values = set(body_columns[position])
        if name.endswith(THRESHOLD_SUFFIXES) and len(values) == 1:
            value = values.pop()
            constants.append((name, value))
            THRESHOLDS[name] = _readable(name, value)
            THRESHOLD_NUMBERS[name] = value
            continue
        keep.append(position)
    header = [header[i] for i in keep]
    body_columns = [body_columns[i] for i in keep]
    body = [[cast(column[row]) for column in body_columns] for row in range(len(raw_body))]

    # blocking_reason: the first dimension that actually excluded this person.
    verdict_at = {name: header.index(name) for name in VERDICTS_IN_ORDER if name in header}
    reasons = []
    for row in body:
        reasons.append(
            next(
                (
                    BLOCKING_REASONS[name]
                    for name, at in verdict_at.items()
                    if row[at] == "FILTERED"
                ),
                None,
            )
            if verdict_at
            else None
        )
    at = header.index("recommend_for_reviewer") + 1 if "recommend_for_reviewer" in header else 0
    header.insert(at, "blocking_reason")
    for row, reason in zip(body, reasons, strict=True):
        row.insert(at, reason)

    # Verdicts move into plain words before anything is styled or written.
    for row in body:
        for position, value in enumerate(row):
            if isinstance(value, str) and value in VERDICT_TEXT:
                row[position] = VERDICT_TEXT[value]

    # A rule that measures something shows the measurement. The verdict it
    # replaces is kept alongside so the cell can still be coloured by it.
    verdicts: list[list[str | None]] = [[None] * len(header) for _ in body]
    for rule, measure in RULE_MEASURES.items():
        if rule not in header or measure not in header:
            continue
        rule_at, measure_at = header.index(rule), header.index(measure)
        for row, shadow in zip(body, verdicts, strict=True):
            if row[measure_at] is None:
                continue  # nothing measured — the word is all there is to say
            shadow[rule_at] = row[rule_at]
            row[rule_at] = row[measure_at]

    # The restricted list is policy, not something this script gets to know. Name
    # it from the journal's own configuration, and cross-check the CSV against it
    # so a policy that has moved on cannot leave a workbook claiming the old rule.
    restricted = restricted_countries(src, journal)
    THRESHOLDS["restricted_countries"] = describe_countries(restricted)
    if restricted and "banned_country_is_banned" in header:
        position = header.index("banned_country_is_banned")
        country = header.index("current_country")
        flagged = {row[country] for row in body if row[position] == 1}
        present = {row[country] for row in body if row[country] in restricted}
        if disagreement := flagged.symmetric_difference(present):
            raise SystemExit(
                "the CSV's restricted-country outcome disagrees with the policy on: "
                + ", ".join(sorted(disagreement))
            )

    if missing := missing_thresholds(header):
        raise SystemExit(
            "the CSV supplies no constant value for: "
            + ", ".join(missing)
            + " — a heading needs them to state its rule"
        )

    wb = Workbook()

    decision = wb.active
    decision.title = "decision"
    picked = [header.index(name) for name in DECISION_COLUMNS if name in header]
    decision_header = [header[i] for i in picked]
    decision_body = [[row[i] for i in picked] for row in body]
    decision_verdicts = [[shadow[i] for i in picked] for shadow in verdicts]
    decision.append(decision_header)
    for row in decision_body:
        decision.append(row)
    _finish(
        decision, decision_header, decision_body, decision_verdicts, freeze="A2", wide=("reasoning",)
    )

    audit = wb.create_sheet("audit")
    audit.append(header)
    for row in body:
        audit.append(row)
    _finish(audit, header, body, verdicts, freeze="A2", wide=("reasoning",))

    glossary = wb.create_sheet("columns")
    for title, text in HOW_TO_READ:
        glossary.append([title, "", text])
        glossary.cell(row=glossary.max_row, column=1).font = Font(bold=True)
    glossary.append(["dimension", "column", "meaning"])
    for cell in glossary[glossary.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
    for name in header:
        glossary.append([dimension_of(name), label_of(name), GLOSSARY.get(name, "(undocumented)")])
    if constants:
        glossary.append([])
        glossary.append(["thresholds", "", "Same for every candidate this run, so not shown as columns."])
        glossary.cell(row=glossary.max_row, column=1).font = Font(bold=True)
        for name, value in constants:
            glossary.append([dimension_of(name), label_of(name), f"= {value}"])
    for column, width in (("A", 24), ("B", 56), ("C", 100)):
        glossary.column_dimensions[column].width = width
    for row in glossary.iter_rows(min_row=1):
        row[2].alignment = Alignment(vertical="top", wrap_text=True)

    wb.active = 0
    wb.save(dst)
    return len(body), len(decision_header), len(header)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("csv", type=Path, help="a 5-shortlist/contact-list-audit.csv")
    parser.add_argument(
        "--journal",
        default="",
        help="journal slug for the policy behind the run; read from the case profile when omitted",
    )
    args = parser.parse_args(argv)
    dst = args.csv.with_suffix(".xlsx")
    rows, decision, audit = build(args.csv, dst, args.journal)
    print(f"{dst}: {rows} rows — decision {decision} columns, audit {audit} columns")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
